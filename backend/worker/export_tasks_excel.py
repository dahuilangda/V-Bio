"""Asynchronous task-list Excel export.

The browser used to build this workbook client-side (ExcelJS + RDKit-WASM),
which stalls the UI and re-downloads every task row when a project has many
tasks. The export now runs as a Celery task on the CPU worker:

* the API server only receives a compact, pre-rendered row payload and returns
  an export id immediately;
* the worker refreshes each row's runtime state from Celery/Redis/result
  archives so the exported "State" column reflects the server-side truth, not
  the possibly stale row in the browser;
* affinity metrics are re-read from the result archive when the task row
  carries none, so the exported indicator set is complete;
* progress counters live in Redis (`ExportJobStore`) and the finished workbook
  is downloaded from the API server, guaranteeing a complete file transfer.

Heavy deps (openpyxl, rdkit, pillow) are imported lazily inside the task so
workers whose images do not carry them can still import this module (and thus
register the task) without crashing at startup.
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import time
import zipfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from celery.result import AsyncResult

from backend.core import config
from backend.core.celery_app import celery_app
from backend.services.export_job_store import ExportJobStore
from gpu_manager import get_redis_client

LOGGER = logging.getLogger(__name__)

EXPORT_IMAGE_WIDTH_PX = 220
EXPORT_IMAGE_HEIGHT_PX = 132
EXPORT_IMAGE_ROW_HEIGHT_PT = round(EXPORT_IMAGE_HEIGHT_PX * 72 / 96)
EXPORT_ATOM_PLDDT_LIMIT = 500
# Affinity JSON members are a few KB; the cap bounds decompression of a corrupt
# or maliciously large member before json.loads (mirrors SCREENING_JSON_MAX_BYTES).
EXPORT_AFFINITY_MEMBER_MAX_BYTES = 8 * 1024 * 1024
# openpyxl rejects these in cell values AND they corrupt sheet-name XML if left
# in the title — strip them everywhere text enters the workbook.
_EXPORT_ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
# Runtime task ids are Celery UUIDs; anything else would also be interpolated
# into result-archive globs (find_result_archive), so reject foreign shapes.
_EXPORT_TASK_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-_.]{0,118}$")
# Result archive files are named "<uuid>_<suffix>.zip"; a fixed-width UUID
# prefix lets one directory listing serve every per-row lookup as a dict hit.
_ARCHIVE_NAME_UUID_PREFIX = re.compile(r"^([0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12})_.+\.zip$")


class ExportArchiveIndex:
    """One listdir of the results directory replaces the per-task glob scan.

    A 13k-row export used to run find_result_archive (several exists() calls
    plus one glob over the whole directory) for EVERY row — quadratic IO on the
    worker. The index answers "which archive belongs to task X" with dict
    lookups; only the rare multi-archive tie-break stats the matching files.
    Archives created AFTER construction are invisible until note_archive() is
    called (used for tasks observed finishing mid-export).
    """

    _CANDIDATE_SUFFIXES = ("_results.zip", "_affinity_results.zip", "_lead_optimization_results.zip")

    def __init__(self, base_dir: str) -> None:
        self.base_dir = base_dir
        self._by_task: Dict[str, list[str]] = {}
        for name in os.listdir(base_dir):
            match = _ARCHIVE_NAME_UUID_PREFIX.match(name)
            if not match:
                continue
            self._by_task.setdefault(match.group(1).lower(), []).append(name)
        for names in self._by_task.values():
            if len(names) > 1:
                names.sort(
                    key=lambda item: os.path.getmtime(os.path.join(base_dir, item)),
                    reverse=True,
                )

    def find_archive(self, task_id: str) -> Optional[str]:
        """Archive file name for the task, or None. UUID-shaped ids only.

        Candidate priority mirrors find_result_archive: the canonical
        <id>_results.zip family first, then the newest <id>_*.zip — so the
        export reads the same archive every other surface resolves.
        """
        normalized = str(task_id or '').strip().lower()
        names = self._by_task.get(normalized)
        if not names:
            return None
        for suffix in self._CANDIDATE_SUFFIXES:
            candidate = f"{normalized}{suffix}"
            if candidate in names:
                return candidate
        return names[0]

    def resolve_path(self, task_id: str) -> Optional[str]:
        name = self.find_archive(task_id)
        return os.path.join(self.base_dir, name) if name else None

    def note_archive(self, task_id: str, name: str) -> None:
        """Register an archive that appeared after construction (rare)."""
        normalized = str(task_id or '').strip().lower()
        if not normalized or normalized in self._by_task:
            return
        self._by_task[normalized] = [name]
EXPORT_TEXT_FIELD_LIMITS = {
    "row_id": 120,
    "task_id": 120,
    "name": 240,
    "summary": 600,
    "backend_label": 60,
    "submitted_text": 60,
    "duration_text": 40,
    "smiles": 4000,
    "interface_label": 12,
}
AFFINITY_FLOAT_FIELDS = (
    "affinity_pred_value",
    "affinity_pic50",
    "affinity_pred_value_mw",
    "affinity_pic50_mw",
    "affinity_pic501",
    "affinity_pic502",
    "affinity_probability_binary",
    "ligand_mw",
)

EXCEL_COLUMNS: Tuple[Tuple[str, float], ...] = (
    ("#", 6),
    ("Task Name", 24),
    ("Task Summary", 36),
    ("Task Row ID", 38),
    ("Runtime Task ID", 24),
    ("State", 12),
    ("Backend", 12),
    ("Submitted", 22),
    ("Duration", 12),
    ("pLDDT", 10),
    ("Interface", 12),
    ("Interface Type", 13),
    ("PAE", 10),
    ("Affinity pIC50", 13),
    ("Affinity pIC50 (MW-corr)", 20),
    ("Affinity log10(IC50 µM)", 20),
    ("Affinity log10(IC50 µM, MW-corr)", 28),
    ("Binder Probability", 16),
    ("Ensemble pIC50 1", 15),
    ("Ensemble pIC50 2", 15),
    ("Ligand MW", 10),
    ("SMILES", 42),
    ("Ligand 2D (Confidence Color)", 36),
)
IMAGE_COLUMN_INDEX = len(EXCEL_COLUMNS)  # 1-based worksheet column of the image


class _UploadFolderShim:
    """Minimal Flask-app stand-in so ResultArchiveService can run inside a worker."""

    def __init__(self, folder: str) -> None:
        self.config = {"UPLOAD_FOLDER": folder}


_archive_service = None


def _get_archive_service():
    global _archive_service
    if _archive_service is None:
        from backend.services.result_archive import ResultArchiveService

        _archive_service = ResultArchiveService(
            app=_UploadFolderShim(config.RESULTS_BASE_DIR),
            celery_app=celery_app,
            logger=LOGGER,
            get_redis_client_fn=get_redis_client,
        )
    return _archive_service


def _to_float(value: Any) -> Optional[float]:
    try:
        if value is None or isinstance(value, bool):
            return None
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number or number in (float("inf"), float("-inf")):
        return None
    return number


def _clean_text(value: Any, limit: int) -> str:
    text = "" if value is None else str(value)
    if len(text) > limit:
        text = text[:limit]
    return _EXPORT_ILLEGAL_XLSX_CHARS.sub("", text).strip()


def normalize_export_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    """Whitelist and bound the client-provided export payload."""
    if not isinstance(payload, dict):
        raise ValueError("Export payload must be a JSON object.")
    raw_rows = payload.get("tasks")
    if not isinstance(raw_rows, list) or not raw_rows:
        raise ValueError("Export payload must contain a non-empty 'tasks' list.")
    max_rows = int(getattr(config, "EXPORT_MAX_TASK_ROWS", 50000))
    if len(raw_rows) > max_rows:
        raise ValueError(f"Export payload exceeds the maximum of {max_rows} task rows.")

    project_name = _clean_text(payload.get("project_name"), 120) or "Tasks"

    rows: List[Dict[str, Any]] = []
    for raw_row in raw_rows:
        if not isinstance(raw_row, dict):
            continue
        row: Dict[str, Any] = {}
        for field, limit in EXPORT_TEXT_FIELD_LIMITS.items():
            row[field] = _clean_text(raw_row.get(field), limit)
        # A task_id containing glob metacharacters would make the result-archive
        # lookup (f"{task_id}_*.zip") match foreign archives; blank it instead so
        # the row degrades to "Draft" rather than importing another task's data.
        if row["task_id"] and not _EXPORT_TASK_ID_PATTERN.match(row["task_id"]):
            row["task_id"] = ""
        if row["row_id"] and not _EXPORT_TASK_ID_PATTERN.match(row["row_id"]):
            row["row_id"] = ""
        metrics_raw = raw_row.get("metrics") if isinstance(raw_row.get("metrics"), dict) else {}
        row["metrics"] = {
            "plddt": _to_float(metrics_raw.get("plddt")),
            "interface_value": _to_float(metrics_raw.get("interface_value")),
            "pae": _to_float(metrics_raw.get("pae")),
        }
        atom_plddts_raw = raw_row.get("atom_plddts")
        atom_plddts: List[float] = []
        if isinstance(atom_plddts_raw, list):
            for item in atom_plddts_raw[:EXPORT_ATOM_PLDDT_LIMIT]:
                number = _to_float(item)
                if number is not None:
                    atom_plddts.append(round(max(0.0, min(100.0, number)), 1))
        row["atom_plddts"] = atom_plddts
        affinity_raw = raw_row.get("affinity") if isinstance(raw_row.get("affinity"), dict) else {}
        row["affinity"] = {
            field: _to_float(affinity_raw.get(field)) for field in AFFINITY_FLOAT_FIELDS
        }
        row["include_image"] = bool(raw_row.get("include_image", True))
        rows.append(row)

    if not rows:
        raise ValueError("Export payload contains no valid task rows.")
    return {
        "export_id": _clean_text(payload.get("export_id"), 80),
        "project_name": project_name,
        "tasks": rows,
    }


def _refresh_runtime_state(task_id: str, archive_index: "ExportArchiveIndex") -> Tuple[str, str]:
    """Authoritative server-side state for a runtime task, mirroring /status logic.

    Returns a (state_token, note) tuple where state_token is one of
    QUEUED/RUNNING/SUCCESS/FAILURE/REVOKED ("" when unknown).
    """
    if not task_id:
        return "", ""
    archive_service = _get_archive_service()
    try:
        result = AsyncResult(task_id, app=celery_app)
        state = str(result.state or "")
        info = result.info
    except Exception as exc:
        LOGGER.warning("Failed to query Celery state for %s: %s", task_id, exc)
        state = "PENDING"
        info = None

    if state == "SUCCESS":
        return "SUCCESS", "Task completed successfully."
    if state == "FAILURE":
        return "FAILURE", "Task failed."
    if state == "REVOKED":
        return "REVOKED", "Task was revoked."

    archive_name = archive_index.find_archive(task_id)
    if archive_name:
        return "SUCCESS", f"Result file found on server ({archive_name})."

    try:
        tracker_status, heartbeat = archive_service.get_tracker_status(task_id)
    except Exception:
        tracker_status, heartbeat = None, None
    if isinstance(tracker_status, dict):
        token = str(tracker_status.get("status") or "").strip().lower()
        if token in {"completed", "complete", "success", "succeeded"}:
            return "SUCCESS", str(tracker_status.get("details") or "Task completed.")
        if token in {"failed", "failure", "timeout", "error"}:
            return "FAILURE", str(tracker_status.get("details") or "Task failed.")
        if token:
            return "RUNNING", str(tracker_status.get("details") or "Task is running.")

    # Mirror /status: a live heartbeat (task_status key expired but the worker is
    # still beating) proves the task is running, not queued.
    if state == "PENDING" and not heartbeat:
        return "QUEUED", "Task is waiting in the queue."

    # Mirror /status's failure-hint scan for non-terminal states: PROGRESS/STARTED
    # meta carrying failure wording must not be exported as "Running".
    if isinstance(info, dict):
        candidate_message = str(
            info.get("status") or info.get("message") or ""
        ).strip()
        lowered = candidate_message.lower()
        if lowered and any(hint in lowered for hint in ("failed", "error", "timeout", "terminated")):
            return "FAILURE", candidate_message or "Task failed."

    return "RUNNING", "Task is running."


def _state_excel_label(state_token: str) -> str:
    return {
        "QUEUED": "Queued",
        "RUNNING": "Running",
        "SUCCESS": "Success",
        "FAILURE": "Failed",
        "REVOKED": "Revoked",
    }.get(state_token, "Draft")


def _pick_affinity_member(names: List[str]) -> Optional[str]:
    """Pick the affinity JSON member from a result zip, mirroring the view-archive builder."""
    lower_names = [name.lower() for name in names]
    if "nesso/manifest.json" in lower_names:
        canonical = next((name for name in names if name.lower() == "nesso/affinity.json"), None)
        if canonical:
            return canonical
        candidates = [name for name in names if name.lower().endswith("/affinity.json")]
        if candidates:
            return sorted(candidates, key=len)[0]
        return None
    protenix = [name for name in names if os.path.basename(name).lower() == "affinity_data.json"]
    if protenix:
        return sorted(protenix, key=len)[0]
    candidates = [
        name
        for name in names
        if name.lower().endswith(".json") and "affinity" in name.lower()
    ]
    if candidates:
        return sorted(candidates, key=lambda item: (len(item), item))[0]
    return None


def _coerce_affinity_dict(payload: Any) -> Dict[str, Any]:
    if isinstance(payload, dict):
        return payload
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                return item
    return {}


def read_affinity_from_archive(task_id: str, archive_index: "ExportArchiveIndex") -> Dict[str, float]:
    """Extract scalar affinity metrics from a task's result archive."""
    if not task_id:
        return {}
    source_zip_path = archive_index.resolve_path(task_id)
    if not source_zip_path:
        return {}
    try:
        with zipfile.ZipFile(source_zip_path, "r") as src_zip:
            names = [name for name in src_zip.namelist() if not name.endswith("/")]
            member = _pick_affinity_member(names)
            if not member:
                return {}
            member_info = src_zip.getinfo(member)
            if member_info.file_size > EXPORT_AFFINITY_MEMBER_MAX_BYTES:
                LOGGER.warning(
                    "Skipping oversized affinity member %s for task %s (%d bytes > cap %d).",
                    member, task_id, member_info.file_size, EXPORT_AFFINITY_MEMBER_MAX_BYTES,
                )
                return {}
            payload = json.loads(src_zip.read(member))
    except Exception as exc:
        LOGGER.debug("Failed to read affinity payload for task %s: %s", task_id, exc)
        return {}
    affinity = _coerce_affinity_dict(payload)
    if not affinity:
        return {}
    # Some backends nest the scalars one level down.
    if not any(field in affinity for field in AFFINITY_FLOAT_FIELDS):
        for nested_key in ("affinity", "result", "prediction", "data"):
            nested = _coerce_affinity_dict(affinity.get(nested_key))
            if nested and any(field in nested for field in AFFINITY_FLOAT_FIELDS):
                affinity = nested
                break
    extracted: Dict[str, float] = {}
    for field in AFFINITY_FLOAT_FIELDS:
        number = _to_float(affinity.get(field))
        if number is not None:
            extracted[field] = number
    return extracted


def _fallback_metrics_from_archive(
    task_id: str, archive_index: "ExportArchiveIndex"
) -> Dict[str, Optional[float]]:
    """Confidence metrics from compact archive extraction (payload-missing fallback)."""
    empty = {"plddt": None, "interface_value": None, "pae": None}
    if not task_id:
        return empty
    source_zip_path = archive_index.resolve_path(task_id)
    if not source_zip_path:
        return empty
    try:
        compact = _get_archive_service().get_compact_prediction_metrics(task_id, source_zip_path=source_zip_path)
    except Exception:
        compact = None
    if not isinstance(compact, dict):
        return empty
    interface = _to_float(compact.get("ligand_ipsae_max"))
    if interface is None:
        interface = _to_float(compact.get("ipsae_dom"))
    if interface is None:
        interface = _to_float(compact.get("pair_iptm"))
    return {
        "plddt": _to_float(compact.get("ligand_plddt")),
        "interface_value": interface,
        "pae": _to_float(compact.get("pair_pae")),
    }


def _confidence_color(value: float) -> Tuple[float, float, float]:
    # AlphaFold confidence palette, matching the client-side renderer.
    if value < 50:
        return (1.0, 0.49, 0.27)
    if value < 70:
        return (1.0, 0.86, 0.07)
    if value < 90:
        return (0.40, 0.80, 0.95)
    return (0.16, 0.47, 0.9)


_RDKIT_DRAWING_AVAILABLE: Optional[bool] = None


def _rdkit_drawing_available() -> bool:
    """RDKit's Cairo drawer needs system X libraries (libXrender/libXext/libX11).

    Slim worker images may lack them; ligand images are cosmetic, so their
    absence must degrade to image-less exports instead of failing the workbook.
    Importing the module is not enough — the backend only breaks at drawer
    CONSTRUCTION, so probe by instantiating a 1x1 Cairo drawer once.
    """
    global _RDKIT_DRAWING_AVAILABLE
    if _RDKIT_DRAWING_AVAILABLE is None:
        try:
            from rdkit.Chem.Draw import rdMolDraw2D

            rdMolDraw2D.MolDraw2DCairo(1, 1)
            _RDKIT_DRAWING_AVAILABLE = True
        except Exception as exc:
            LOGGER.warning(
                "RDKit 2D drawing is unavailable on this worker (%s); "
                "Excel exports will omit ligand images.",
                exc,
            )
            _RDKIT_DRAWING_AVAILABLE = False
    return _RDKIT_DRAWING_AVAILABLE


def render_ligand_2d_png(smiles: str, atom_plddts: List[float]) -> Optional[bytes]:
    """Render a 2D ligand depiction colored by per-atom confidence (RDKit Cairo)."""
    if not _rdkit_drawing_available():
        return None
    from rdkit import Chem
    from rdkit.Chem.Draw import rdMolDraw2D

    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return None
    highlight_atoms: List[int] = []
    highlight_colors: Dict[int, Tuple[float, float, float]] = {}
    highlight_radii: Dict[int, float] = {}
    atom_count = mol.GetNumAtoms()
    if atom_plddts and len(atom_plddts) == atom_count:
        radius = 0.21 if atom_count > 90 else 0.24 if atom_count > 55 else 0.28
        for index, confidence in enumerate(atom_plddts):
            highlight_atoms.append(index)
            highlight_colors[index] = _confidence_color(float(confidence))
            highlight_radii[index] = radius
    try:
        drawer = rdMolDraw2D.MolDraw2DCairo(EXPORT_IMAGE_WIDTH_PX, EXPORT_IMAGE_HEIGHT_PX)
        drawer.drawOptions().clearBackground = True
        if highlight_atoms:
            drawer.DrawMolecule(
                mol,
                highlightAtoms=highlight_atoms,
                highlightAtomColors=highlight_colors,
                highlightAtomRadii=highlight_radii,
                highlightBonds=[],
            )
        else:
            drawer.DrawMolecule(mol)
        drawer.FinishDrawing()
        return drawer.GetDrawingText()
    except Exception as exc:
        LOGGER.debug("Failed to render ligand 2D image for %s...: %s", smiles[:40], exc)
        return None


def _format_metric(value: Optional[float], digits: int) -> str:
    if value is None:
        return "-"
    return f"{value:.{digits}f}"


def _format_interface_label(raw: Any) -> str:
    """Normalize the interface metric label, preserving the ipTM casing."""
    token = str(raw or "").strip().lower()
    return {"ipsae": "IPSAE", "iptm": "ipTM"}.get(token, token.upper())


def _sanitize_file_stem(value: str) -> str:
    cleaned = re.sub(r"[^\w\-.]+", "_", value, flags=re.UNICODE).strip("._-")
    return cleaned[:60] or "tasks"


def _cleanup_expired_export_files(exports_dir: str, ttl_seconds: int) -> None:
    now = time.time()
    try:
        entries = os.listdir(exports_dir)
    except Exception:
        return
    for entry in entries:
        if not entry.endswith(".xlsx"):
            continue
        path = os.path.join(exports_dir, entry)
        try:
            if now - os.path.getmtime(path) > ttl_seconds:
                os.remove(path)
        except Exception as exc:
            LOGGER.debug("Skipping expired export cleanup for %s: %s", path, exc)


def build_tasks_excel_workbook(
    project_name: str,
    rows: List[Dict[str, Any]],
    *,
    archive_index: "ExportArchiveIndex",
    progress_callback=None,
) -> Tuple[bytes, int, int]:
    """Build the workbook; returns (xlsx_bytes, rendered_image_count, smiles_row_count)."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet_name = (
        _EXPORT_ILLEGAL_XLSX_CHARS.sub("", re.sub(r"[\\/*?:\[\]]+", " ", project_name)).strip()[:31]
        or "Tasks"
    )
    worksheet = workbook.active
    worksheet.title = sheet_name

    for column_index, (header, width) in enumerate(EXCEL_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    rendered_images = 0
    smiles_rows = 0
    image_column_letter = get_column_letter(IMAGE_COLUMN_INDEX)

    for index, row in enumerate(rows):
        task_id = row.get("task_id") or ""
        state_token, state_note = _refresh_runtime_state(task_id, archive_index)
        state_label = _state_excel_label(state_token)
        if state_token == "FAILURE" and state_note:
            LOGGER.debug("Export row %s (task %s) state FAILURE: %s", row.get("row_id"), task_id, state_note)
        if task_id and state_token == "SUCCESS" and not archive_index.find_archive(task_id):
            # Task finished while this export was running — its archive
            # postdates the index snapshot. One targeted lookup (rare) keeps
            # enrichment correct for exactly these rows.
            late_archive = _get_archive_service().find_result_archive(task_id)
            if late_archive:
                archive_index.note_archive(task_id, late_archive)

        # Enrich from the result archive whenever ANY indicator is missing — the
        # archive is the authoritative source; payload values only fill gaps.
        affinity: Dict[str, float] = dict(row.get("affinity") or {})
        if task_id and any(affinity.get(field) is None for field in AFFINITY_FLOAT_FIELDS):
            for field, value in read_affinity_from_archive(task_id, archive_index).items():
                if value is not None:
                    affinity[field] = value

        metrics = dict(row.get("metrics") or {})
        if task_id and any(metrics.get(key) is None for key in ("plddt", "interface_value", "pae")):
            fallback = _fallback_metrics_from_archive(task_id, archive_index)
            for key in ("plddt", "interface_value", "pae"):
                if metrics.get(key) is None and fallback.get(key) is not None:
                    metrics[key] = fallback[key]

        image_bytes: Optional[bytes] = None
        smiles = row.get("smiles") or ""
        if smiles and row.get("include_image", True):
            smiles_rows += 1
            image_bytes = render_ligand_2d_png(smiles, list(row.get("atom_plddts") or []))
            if image_bytes is not None:
                rendered_images += 1

        excel_row_number = index + 2
        values = [
            index + 1,
            row.get("name") or "",
            row.get("summary") or "",
            row.get("row_id") or "",
            task_id,
            state_label,
            row.get("backend_label") or "",
            row.get("submitted_text") or "",
            row.get("duration_text") or "",
            _format_metric(metrics.get("plddt"), 1),
            _format_metric(metrics.get("interface_value"), 3),
            _format_interface_label(row.get("interface_label")),
            _format_metric(metrics.get("pae"), 2),
            _format_metric(affinity.get("affinity_pic50"), 2),
            _format_metric(affinity.get("affinity_pic50_mw"), 2),
            _format_metric(affinity.get("affinity_pred_value"), 3),
            _format_metric(affinity.get("affinity_pred_value_mw"), 3),
            _format_metric(affinity.get("affinity_probability_binary"), 3),
            _format_metric(affinity.get("affinity_pic501"), 2),
            _format_metric(affinity.get("affinity_pic502"), 2),
            _format_metric(affinity.get("ligand_mw"), 2),
            smiles,
            "-" if image_bytes is None and smiles else "",
        ]
        for column_index, value in enumerate(values, start=1):
            if isinstance(value, str) and _EXPORT_ILLEGAL_XLSX_CHARS.search(value):
                # Defense in depth: normalization already strips these, but a direct
                # caller must not be able to abort the whole workbook with one cell.
                value = _EXPORT_ILLEGAL_XLSX_CHARS.sub("", value)
            worksheet.cell(row=excel_row_number, column=column_index, value=value)

        if image_bytes is not None:
            worksheet.row_dimensions[excel_row_number].height = EXPORT_IMAGE_ROW_HEIGHT_PT
            image = XLImage(io.BytesIO(image_bytes))
            image.width = EXPORT_IMAGE_WIDTH_PX
            image.height = EXPORT_IMAGE_HEIGHT_PX
            worksheet.add_image(image, f"{image_column_letter}{excel_row_number}")

        if progress_callback is not None:
            progress_callback(index + 1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), rendered_images, smiles_rows


@celery_app.task(bind=True, name="tasks.export_tasks_excel_task")
def export_tasks_excel_task(self, payload: dict):
    export_id = str((payload or {}).get("export_id") or "").strip().lower()
    store = ExportJobStore(
        get_redis_client_fn=get_redis_client,
        logger=LOGGER,
        ttl_seconds=int(getattr(config, "EXPORT_JOB_TTL_SECONDS", 48 * 3600)),
    )
    started_at = time.time()
    # The export_id becomes a file path component and a Redis key — only the
    # 32-hex ids the API server generates are acceptable (fail fast otherwise).
    if not re.fullmatch(r"[a-f0-9]{32}", export_id):
        LOGGER.error("Rejected export payload with invalid export_id %r.", export_id)
        return {"status": "failure", "error": "Invalid export_id."}
    try:
        normalized = normalize_export_payload(payload or {})
    except ValueError as exc:
        store.update(export_id, status="failure", error=str(exc))
        LOGGER.warning("Rejected export payload: %s", exc)
        return {"status": "failure", "error": str(exc)}

    try:
        rows = normalized["tasks"]
        os.makedirs(config.EXPORTS_BASE_DIR, exist_ok=True)
        _cleanup_expired_export_files(
            config.EXPORTS_BASE_DIR, int(getattr(config, "EXPORT_FILE_TTL_SECONDS", 48 * 3600))
        )
        # One directory listing serves every per-row archive lookup in this
        # export; construction failure (unreadable results dir) fails the
        # export honestly instead of stranding a running job.
        archive_index = ExportArchiveIndex(config.RESULTS_BASE_DIR)
        # The API server creates the job record before dispatch; self-create
        # only if the record vanished (Redis flush) so status updates never
        # silently no-op against a missing record.
        if store.load(export_id) is None:
            store.create(
                export_id=export_id,
                celery_task_id=str(getattr(self.request, "id", "") or ""),
                project_name=normalized["project_name"],
                total=len(rows),
                queue="",
            )
        store.update(export_id, status="running", total=len(rows), done=0, error="")

        def _report_progress(done: int) -> None:
            store.update(export_id, done=done)

        xlsx_bytes, rendered_images, smiles_rows = build_tasks_excel_workbook(
            normalized["project_name"],
            rows,
            archive_index=archive_index,
            progress_callback=_report_progress,
        )
        warning = ""
        if smiles_rows > 0 and rendered_images == 0:
            # Image rendering is cosmetic; a workbook whose ligands cannot be
            # depicted must still be delivered instead of failing every other row.
            warning = (
                f"Ligand 2D rendering failed for all {smiles_rows} SMILES rows; "
                "the workbook was produced without ligand images."
            )
            LOGGER.warning("Excel export %s: %s", export_id, warning)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_name = f"{_sanitize_file_stem(normalized['project_name'])}_tasks_{timestamp}.xlsx"
        # Atomic publish: write + rename so a concurrent download (or an
        # acks_late redelivery re-running this task) never reads a partial file.
        file_path = os.path.join(config.EXPORTS_BASE_DIR, f"{export_id}.xlsx")
        tmp_path = f"{file_path}.tmp"
        with open(tmp_path, "wb") as handle:
            handle.write(xlsx_bytes)
        os.replace(tmp_path, file_path)
        store.update(
            export_id,
            status="success",
            done=len(rows),
            file_name=file_name,
            file_bytes=len(xlsx_bytes),
            warning=warning,
        )
        LOGGER.info(
            "Excel export %s finished: %d rows, %d images, %.1fs, %s bytes.",
            export_id,
            len(rows),
            rendered_images,
            time.time() - started_at,
            len(xlsx_bytes),
        )
        return {
            "status": "success",
            "export_id": export_id,
            "file_name": file_name,
            "rows": len(rows),
            "images": rendered_images,
        }
    except Exception as exc:
        error_message = f"{type(exc).__name__}: {exc}"
        store.update(export_id, status="failure", error=error_message[:2000])
        LOGGER.exception("Excel export %s failed.", export_id)
        return {"status": "failure", "export_id": export_id, "error": error_message[:2000]}
